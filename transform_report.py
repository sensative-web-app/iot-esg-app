#!/usr/bin/env python3

import sys
from itertools import pairwise, islice
from pprint import pprint
from datetime import datetime, timedelta
from statistics import mean

import openpyxl


def main(input_file, output_file):
    wb = openpyxl.load_workbook(input_file)
    sheet = wb["timeseriesRawData"]

    stuff, rstuff = extract_sheet_data(sheet)
    pprint(stuff)
    pprint(rstuff)

    results_elec = extract_prop_data(
        "Electricity consumption",
        sheet, stuff, rstuff)
    (result_elec,) = results_elec.values()
    #pprint(result_elec)

    elec_con = find_consumption_in_period(result_elec, timedelta(days=7))
    print(elec_con)

    results_water = extract_prop_data(
        "Volume!",
        sheet, stuff, rstuff)
    pprint(results_water.keys())
    it = iter(results_water.values())
    cons = [find_consumption_in_period(thing, timedelta(days=7))
            for thing in results_water.values()]
    #print(cons, sum(cons) / 1000000)

    results_hum = extract_prop_data(
        "Humidity",
        sheet, stuff, rstuff)
    (result_hum,) = results_hum.values()
    index = find_backwards(result_hum, timedelta(days=7))
    hum_points = result_hum[index:]

    hum_values = [point[1] for point in hum_points]
    print(min(hum_values), max(hum_values), mean(hum_values))

    results_temp = extract_prop_data(
        "Air temperature",
        sheet, stuff, rstuff)
    (result_temp,) = results_temp.values()
    index = find_backwards(result_temp, timedelta(days=7))
    temp_points = result_temp[index:]

    temp_values = [point[1] for point in temp_points]
    print(min(temp_values), max(temp_values), mean(temp_values))

    results_co2 = extract_prop_data(
        "CO2",
        sheet, stuff, rstuff)
    (result_co2,) = results_co2.values()
    index = find_backwards(result_co2, timedelta(days=7))
    co2_points = result_co2[index:]

    co2_values = [point[1] for point in co2_points]
    print(min(co2_values), max(co2_values), mean(co2_values))

    outsheet_name = "transformedData"
    if outsheet_name in wb:
        del wb[outsheet_name]
    outsheet = wb.create_sheet(outsheet_name)

    outsheet.append(("Electricity consumption", elec_con))
    outsheet.append(("Water consumption", sum(cons) / 1000000))
    outsheet.append(("Humidity", min(hum_values), max(hum_values), mean(hum_values)))
    outsheet.append(("Air temperature", min(temp_values), max(temp_values), mean(temp_values)))
    outsheet.append(("CO2", min(co2_values), max(co2_values), mean(co2_values)))
    wb.save(output_file)


def extract_sheet_data(sheet):
    properties = {"Air temperature", "CO2", "Humidity", "Volume!",
                  "Electricity consumption"}
    start_cols = [cell.col_idx for cell in sheet[1]
                  if not cell.value in properties]

    stuff = {}

    for start_col, end_col in pairwise(start_cols + [len(sheet[1]) + 1]):
        cell = sheet[1][start_col - 1]
        name = cell.value
        name_pos = cell.col_idx
        props = [(cell.value, cell.col_idx) for cell in
                 [sheet[1][col-1] for col in range(start_col + 1, end_col)]]
        stuff[name] = (name_pos, props)

    rstuff = {}
    for name, (name_pos, props) in stuff.items():
        for prop_name, prop_pos in props:
            #print(prop_name)
            rstuff.setdefault(prop_name, [])
            rstuff[prop_name].append((name, prop_pos))

    return stuff, rstuff


def extract_prop_data(key, sheet, stuff, rstuff):
    results = {}
    #result = []
    for name, prop_pos in rstuff[key]:
        it = iter(sheet.iter_rows())
        next(it)  # skip headings
        #it = islice(it, 7)
        results.setdefault(name, [])
        result = results[name]
        for row in it:
            name_pos = stuff[name][0]
            value = row[prop_pos - 1].value
            ts = row[name_pos - 1].value
            if value is None or ts is None:
                continue
            #print([cell.value for cell in row])
            result.append((ts, value, name))
            #print(ts, name_pos, prop_pos, value)
            #print(result)
    return results


def find_backwards(data, delta):
    last_ts = parse_ts(data[-1][0])
    target_ts = last_ts - delta
    found_index = None
    for i, row in enumerate(reversed(data)):
        ts = row[0]
        ts = parse_ts(ts)
        if ts < target_ts:
            found_index = len(data) - i - 1
            assert parse_ts(data[found_index][0]) == ts
            return found_index
    else:
        return 0


def find_consumption_in_period(data, delta):
    index = find_backwards(data, delta)
    first = data[index]
    last = data[-1]
    return last[1] - first[1]


def parse_ts(s):
    try:
        return datetime.strptime(s, "%Y-%m-%dT%H:%M:%S.%fZ")
    except ValueError:
        return datetime.strptime(s, "%Y-%m-%dT%H:%M:%SZ")


if __name__ == "__main__":
    main(sys.argv[1], sys.argv[2])
